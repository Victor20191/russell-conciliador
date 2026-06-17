import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

with open("/tmp/catalogos.json", encoding="utf-8") as f:
    cat = json.load(f)
usuarios = cat["usuarios"]

FONT = "Calibri"
NAVY = "1F3A5F"
NAVY_SOFT = "33507A"
REQ_FILL = "FFF4E0"
OPT_FILL = "EEF2F7"
EX_FILL = "F5F7FA"
WHITE = "FFFFFF"
thin = Side(style="thin", color="C9D2DE")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def bf(**kw):
    return Font(name=FONT, **{k: v for k, v in kw.items() if k in ("bold", "italic", "size", "color")})

wb = Workbook()

# =================================================================
# HOJA · INSTRUCCIONES
# =================================================================
ws = wb.active
ws.title = "Instrucciones"
ws.sheet_view.showGridLines = False
for col, w in (("A", 3), ("B", 28), ("C", 78), ("D", 16)):
    ws.column_dimensions[col].width = w

def put(cell, value, **kw):
    ws[cell] = value
    if kw:
        ws[cell].font = bf(**kw)
    return ws[cell]

ws.merge_cells("B2:D2")
put("B2", "Maestros de personas — Socio · Gerente · Senior · Staff", bold=True, size=16, color=NAVY)
ws.merge_cells("B3:D3")
put("B3", "Russell Bedford · Diagnóstico — cada persona es un usuario del sistema", italic=True, size=10, color="6B7585")

ws.merge_cells("B5:D5")
put("B5", "Cómo usar esta plantilla", bold=True, size=12, color=NAVY)
pasos = [
    "1.  Hay una hoja por tipo: «Socios», «Gerentes», «Seniors» y «Staff». El tipo (rol) lo define la hoja, no se captura.",
    "2.  Captura una persona por fila. Columnas con * son obligatorias (fondo ámbar).",
    "3.  Campos clave de cada persona: Nombre completo, Cédula (única) y Cargo (título de puesto, texto libre).",
    "4.  Correo: es la identidad de inicio de sesión (único). La contraseña se genera temporal y se exige cambiarla en el primer ingreso.",
    "5.  ORDEN DE CARGA: primero «Socios», luego «Gerentes», «Seniors» y por último «Staff».",
    "6.  Jerarquía: Gerente referencia la Cédula de su Socio; Senior la de su Gerente; Staff la de su Senior. Esa cédula debe existir (cargada antes o en este mismo archivo).",
    "7.  Estado: «Activo» (por defecto) o «Inactivo».",
    "8.  Borra las filas de EJEMPLO (en gris) antes de importar.",
]
r = 6
for p in pasos:
    ws.merge_cells(f"B{r}:D{r}")
    put(f"B{r}", p, size=10.5)
    ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 30
    r += 1

# --- Modelo de datos / relaciones ---
r += 1
ws.merge_cells(f"B{r}:D{r}")
put(f"B{r}", "Cómo se guardan (tablas y relaciones)", bold=True, size=12, color=NAVY)
r += 1
hdr = r
for col, txt in (("B", "Tabla física"), ("C", "Qué guarda / relación"), ("D", "")):
    c = ws[f"{col}{hdr}"]
    c.value = txt
    c.font = bf(bold=True, color=WHITE, size=10.5)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = border
ws.merge_cells(f"C{hdr}:D{hdr}")
ws.row_dimensions[hdr].height = 20
r += 1
modelo = [
    ("usuarios", "Cada persona es un usuario. Campos: nombre, cedula (único), cargo, correo (único), rol = el tipo de la hoja, activo."),
    ("jerarquia_usuarios", "Relación superior→subordinado: Socio→Gerente→Senior→Staff. Se arma con la «Cédula del superior» de cada hoja."),
    ("asignaciones_cliente", "Quién es responsable de cada cliente (Staff ejecuta, Senior/Gerente consultan). Apunta a usuarios.id."),
    ("clientes.socio_id", "Socio (firma) del cliente — referencia informativa a usuarios.id (no otorga acceso)."),
]
for tabla, desc in modelo:
    ws.merge_cells(f"C{r}:D{r}")
    ws[f"B{r}"].value = tabla
    ws[f"B{r}"].font = bf(bold=True, size=10)
    ws[f"C{r}"].value = desc
    ws[f"C{r}"].font = bf(size=10)
    ws[f"C{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    for col in ("B", "C", "D"):
        ws[f"{col}{r}"].border = border
    ws.row_dimensions[r].height = 32
    r += 1

# --- Usuarios ya existentes (para no duplicar correo/cédula) ---
r += 1
ws.merge_cells(f"B{r}:D{r}")
put(f"B{r}", "Usuarios ya existentes (no repitas correo ni cédula)", bold=True, size=12, color=NAVY)
r += 1
for col, txt in (("B", "Nombre"), ("C", "Correo"), ("D", "Rol")):
    c = ws[f"{col}{r}"]
    c.value = txt
    c.font = bf(bold=True, color=WHITE, size=10.5)
    c.fill = PatternFill("solid", fgColor=NAVY_SOFT)
    c.border = border
ws.row_dimensions[r].height = 18
r += 1
for u in sorted(usuarios, key=lambda x: (x["role"], x["name"])):
    ws[f"B{r}"].value = u["name"]
    ws[f"C{r}"].value = u["email"]
    ws[f"D{r}"].value = u["role"]
    for col in ("B", "C", "D"):
        ws[f"{col}{r}"].font = bf(size=9.5)
        ws[f"{col}{r}"].border = border
    r += 1

# =================================================================
# Hoja de LISTAS (desplegables)
# =================================================================
wl = wb.create_sheet("Listas")
wl.sheet_view.showGridLines = False
wl["A1"] = "Estado"
wl["A1"].font = bf(bold=True, color=WHITE, size=10.5)
wl["A1"].fill = PatternFill("solid", fgColor=NAVY)
wl["A1"].alignment = Alignment(horizontal="center")
wl.column_dimensions["A"].width = 14
for i, v in enumerate(["Activo", "Inactivo"], start=2):
    wl[f"A{i}"] = v
    wl[f"A{i}"].font = bf(size=10)
ESTADO_RNG = "Listas!$A$2:$A$3"

# =================================================================
# Hojas de MAESTROS
# =================================================================
def crear_maestro(nombre_hoja, rol, superior_label, cargo_ejemplo, ejemplos):
    ws = wb.create_sheet(nombre_hoja)
    ws.sheet_view.showGridLines = False
    headers = [
        ("Nombre completo *", 30, True),
        ("Cédula *", 18, True),
        ("Cargo *", 26, True),
        ("Correo *", 30, True),
    ]
    if superior_label:
        headers.append((superior_label, 26, True))
    headers.append(("Estado", 14, False))

    # Banda de título con el rol asignado
    last_col = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"Maestro de {nombre_hoja} · rol asignado: {rol}"
    ws["A1"].font = bf(bold=True, color=WHITE, size=11)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24

    for idx, (title, width, req) in enumerate(headers, start=1):
        col = get_column_letter(idx)
        ws.column_dimensions[col].width = width
        c = ws[f"{col}2"]
        c.value = title
        c.font = bf(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=(NAVY_SOFT if req else "5A6B82"))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"
    n_cols = len(headers)
    n_req = n_cols - 1  # todas menos Estado

    # Filas de ejemplo
    for ri, fila in enumerate(ejemplos, start=3):
        for ci, val in enumerate(fila, start=1):
            cell = ws[f"{get_column_letter(ci)}{ri}"]
            cell.value = val
            cell.font = bf(size=10, italic=True, color="8A94A3")
            cell.fill = PatternFill("solid", fgColor=EX_FILL)
            cell.border = border
        nota = ws[f"{get_column_letter(n_cols + 2)}{ri}"]
        nota.value = "← EJEMPLO (borrar)"
        nota.font = bf(size=9, italic=True, bold=True, color="B22222")

    first_data = 3 + len(ejemplos)
    last_row = 300
    for ri in range(first_data, last_row + 1):
        for ci in range(1, n_cols + 1):
            cell = ws[f"{get_column_letter(ci)}{ri}"]
            cell.border = border
            cell.font = bf(size=10)
            cell.fill = PatternFill("solid", fgColor=(REQ_FILL if ci <= n_req else OPT_FILL))

    # Validación de Estado (última columna)
    dv = DataValidation(type="list", formula1=ESTADO_RNG, allow_blank=True,
                        showErrorMessage=True, errorStyle="stop",
                        showInputMessage=True)
    dv.prompt = "Activo (por defecto) o Inactivo"
    dv.promptTitle = "Estado"
    dv.error = "Elige Activo o Inactivo."
    dv.errorTitle = "Valor no válido"
    estado_col = get_column_letter(n_cols)
    dv.add(f"{estado_col}3:{estado_col}{last_row}")
    ws.add_data_validation(dv)
    return ws

crear_maestro(
    "Socios", "Socio", None, "Socio Director",
    [["Carlos Andrés Gómez", "79.123.456", "Socio Director", "cgomez@russellbedford.co", "Activo"],
     ["María Fernanda Ríos", "52.987.654", "Socia de Impuestos", "mrios@russellbedford.co", "Activo"]],
)
crear_maestro(
    "Gerentes", "Gerente", "Cédula del Socio *", "Gerente de Auditoría",
    [["Juan Pablo Mejía", "1.020.304.050", "Gerente de Auditoría", "jmejia@russellbedford.co", "79.123.456", "Activo"]],
)
crear_maestro(
    "Seniors", "Senior", "Cédula del Gerente *", "Auditor Senior",
    [["Laura Restrepo", "1.130.250.360", "Auditor Senior", "lrestrepo@russellbedford.co", "1.020.304.050", "Activo"]],
)
crear_maestro(
    "Staff", "Staff", "Cédula del Senior *", "Auditor Junior",
    [["Andrés Quintero", "1.040.560.780", "Auditor Junior", "aquintero@russellbedford.co", "1.130.250.360", "Activo"]],
)

# Orden de hojas: Instrucciones, Socios, Gerentes, Seniors, Staff, Listas
orden = ["Instrucciones", "Socios", "Gerentes", "Seniors", "Staff", "Listas"]
wb._sheets.sort(key=lambda s: orden.index(s.title))
wb.active = wb.sheetnames.index("Socios")

out = "/Users/vicbook/Documents/Xentria-apps/Russell Diagnostico/russell-lfm/Plantilla_Maestros_Personas.xlsx"
wb.save(out)
print("OK", out)
