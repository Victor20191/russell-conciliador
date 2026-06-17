import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

with open("/tmp/catalogos.json", encoding="utf-8") as f:
    cat = json.load(f)

usuarios = cat["usuarios"]
aristas = cat["aristas"]
by_id = {u["id"]: u for u in usuarios}

def activos_de(rol):
    return sorted([u["name"] for u in usuarios if u["active"] and u["role"] == rol])

socios = activos_de("Socio")
gerentes = activos_de("Gerente")
seniors = activos_de("Senior")
staffs = activos_de("Staff")
erps = cat["erps"]
sectores = cat["sectors"]
modulos = [m["name"] for m in cat["modules"]]            # orden alfabético del catálogo
dian = [(f["name"], f["code"]) for f in cat["dianForms"]]
tipos = ["A", "B", "C"]

# Cadenas válidas de jerarquía: socio -> gerente -> senior -> [staff]
nombre = lambda i: by_id[i]["name"]
sup_de = {}
for a in aristas:
    sup_de.setdefault(a["superiorId"], []).append(a["subordinateId"])

# El Socio es informativo (independiente): no entra en la cascada. La cascada
# que el formulario valida es Gerente → Senior → Staff.
cadenas = []  # (gerente, senior, [staff...])
for g in usuarios:
    if not (g["active"] and g["role"] == "Gerente"):
        continue
    for sid in sup_de.get(g["id"], []):
        s = by_id.get(sid)
        if not (s and s["active"] and s["role"] == "Senior"):
            continue
        st = [nombre(x) for x in sup_de.get(sid, [])
              if by_id.get(x) and by_id[x]["active"] and by_id[x]["role"] == "Staff"]
        cadenas.append((g["name"], s["name"], sorted(st)))

# ---------- Estilos ----------
FONT = "Calibri"
NAVY = "1F3A5F"
NAVY_SOFT = "33507A"
REQ_FILL = "FFF4E0"      # ámbar suave para requeridos
OPT_FILL = "EEF2F7"      # gris azulado para opcionales
EX_FILL = "F5F7FA"       # filas de ejemplo
WHITE = "FFFFFF"
thin = Side(style="thin", color="C9D2DE")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def base_font(**kw):
    return Font(name=FONT, **{k: v for k, v in kw.items()
                              if k in ("bold", "italic", "size", "color")})

wb = Workbook()

# =================================================================
# HOJA 1 · INSTRUCCIONES
# =================================================================
ws = wb.active
ws.title = "Instrucciones"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 70
ws.column_dimensions["D"].width = 26
ws.column_dimensions["E"].width = 30

def put(cell, value, **kw):
    ws[cell] = value
    if kw:
        ws[cell].font = base_font(**kw)
    return ws[cell]

ws.merge_cells("B2:E2")
put("B2", "Plantilla de importación masiva de clientes", bold=True, size=16, color=NAVY)
ws.merge_cells("B3:E3")
put("B3", "Russell Bedford · Diagnóstico — ruta /config/clientes", italic=True, size=10, color="6B7585")

ws.merge_cells("B5:E5")
put("B5", "Cómo usar esta plantilla", bold=True, size=12, color=NAVY)

pasos = [
    "1.  Captura un cliente por fila en la hoja «Clientes». No cambies los encabezados ni el orden de las columnas.",
    "2.  Las columnas marcadas con * son obligatorias (fondo ámbar). El resto es opcional (fondo gris).",
    "3.  El «Código» (C-####) NO se captura: la plataforma lo asigna automáticamente al importar.",
    "4.  «Tipo de cliente» es A, B o C (lista desplegable).",
    "5.  ERP, Sector, Socio, Gerente y Senior tienen lista desplegable. Para módulos y DIAN elige «Sí» o «No» (vacío = No).",
    "6.  «Staff (ejecuta)» admite uno o varios: escribe los nombres separados por punto y coma «;».",
    "7.  El Socio (firma) se registra como dato informativo: elige cualquier Socio activo (no condiciona al Gerente).",
    "8.  Cascada de equipo: el Senior debe reportar al Gerente y cada Staff al Senior (ver tabla abajo).",
    "9.  Borra las filas de EJEMPLO (en gris) antes de importar.",
]
r = 6
for p in pasos:
    ws.merge_cells(f"B{r}:E{r}")
    put(f"B{r}", p, size=10.5)
    ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 30
    r += 1

# --- Tabla de referencia de columnas ---
r += 1
ws.merge_cells(f"B{r}:E{r}")
put(f"B{r}", "Referencia de columnas", bold=True, size=12, color=NAVY)
r += 1
hdr = r
ws.merge_cells(f"C{hdr}:D{hdr}")
for col, txt in (("B", "Columna"), ("C", "Qué escribir / reglas"), ("E", "Obligatorio")):
    c = ws[f"{col}{hdr}"]
    c.value = txt
    c.font = base_font(bold=True, color=WHITE, size=10.5)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
ws[f"C{hdr}"].border = border
ws[f"B{hdr}"].border = border
ws[f"E{hdr}"].border = border
ws.row_dimensions[hdr].height = 20
r += 1

ref = [
    ("Razón social", "Nombre legal del cliente. Texto libre.", "Sí"),
    ("NIT", "NIT con o sin dígito de verificación (ej. 900.451.227-3).", "Sí"),
    ("Tipo de cliente", "Clasificación del cliente: A, B o C.", "Sí"),
    ("ERP", "Sistema contable. Elige de la lista o escribe uno nuevo.", "Sí"),
    ("Sector", "Sector económico. Elige de la lista o escribe uno nuevo.", "Sí"),
    ("Socio (firma)", "Un Socio activo (informativo). Solo de la lista desplegable.", "Sí"),
    ("Gerente (valida)", "Un Gerente activo. Solo de la lista desplegable.", "Sí"),
    ("Senior (revisa)", "Un Senior activo que reporte al Gerente elegido.", "Sí"),
    ("Staff (ejecuta)", "Uno o varios Staff que reporten al Senior, separados por «;».", "Sí"),
    ("Módulos (6 columnas)", "Activos fijos, Cartera, Cuentas por pagar, Ingresos, Inventarios, Nómina. Sí/No.", "No"),
    ("DIAN (3 columnas)", "IVA (F-300), Retención en la fuente (F-350), ICA (F-CHIP). Sí/No.", "No"),
]
for nombre_col, regla, obl in ref:
    ws.merge_cells(f"C{r}:D{r}")
    ws[f"B{r}"].value = nombre_col
    ws[f"B{r}"].font = base_font(bold=True, size=10)
    ws[f"C{r}"].value = regla
    ws[f"C{r}"].font = base_font(size=10)
    ws[f"C{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws[f"E{r}"].value = obl
    ws[f"E{r}"].font = base_font(size=10, bold=(obl == "Sí"), color=("B25E00" if obl == "Sí" else "6B7585"))
    ws[f"E{r}"].alignment = Alignment(horizontal="center", vertical="top")
    for col in ("B", "C", "E"):
        ws[f"{col}{r}"].border = border
    ws[f"D{r}"].border = border
    ws.row_dimensions[r].height = 28
    r += 1

# --- Socios disponibles (informativo, independiente de la cascada) ---
r += 1
ws.merge_cells(f"B{r}:E{r}")
put(f"B{r}", "Socios disponibles (firma · informativo)", bold=True, size=12, color=NAVY)
r += 1
ws.merge_cells(f"B{r}:E{r}")
put(f"B{r}", "; ".join(socios) if socios else "No hay socios activos.", size=10.5)
ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r].height = 24
r += 1

# --- Cascada de equipo válida (Gerente → Senior → Staff) ---
r += 1
ws.merge_cells(f"B{r}:E{r}")
put(f"B{r}", "Cascada de equipo válida hoy (Gerente → Senior → Staff)", bold=True, size=12, color=NAVY)
r += 1
ws.merge_cells(f"D{r}:E{r}")
for col, txt in (("B", "Gerente (valida)"), ("C", "Senior (revisa)"),
                 ("D", "Staff disponibles (ejecuta)")):
    c = ws[f"{col}{r}"]
    c.value = txt
    c.font = base_font(bold=True, color=WHITE, size=10.5)
    c.fill = PatternFill("solid", fgColor=NAVY_SOFT)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = border
ws[f"E{r}"].border = border
ws.row_dimensions[r].height = 20
r += 1
if cadenas:
    for g, s, st in cadenas:
        ws.merge_cells(f"D{r}:E{r}")
        ws[f"B{r}"].value = g
        ws[f"C{r}"].value = s
        ws[f"D{r}"].value = "; ".join(st) if st else "(sin staff a cargo)"
        for col in ("B", "C", "D"):
            ws[f"{col}{r}"].font = base_font(size=10)
            ws[f"{col}{r}"].border = border
            ws[f"{col}{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"E{r}"].border = border
        ws.row_dimensions[r].height = 24
        r += 1
else:
    ws.merge_cells(f"B{r}:E{r}")
    put(f"B{r}", "No hay cascadas de equipo completas configuradas todavía.", italic=True, color="B22222")
    r += 1

# =================================================================
# HOJA · LISTAS  (fuente de los desplegables)
# =================================================================
wl = wb.create_sheet("Listas")
listas = {
    "A": ("Sí/No", ["Sí", "No"]),
    "B": ("ERP", erps),
    "C": ("Sector", sectores),
    "D": ("Socio", socios),
    "E": ("Gerente", gerentes),
    "F": ("Senior", seniors),
    "G": ("Staff", staffs),
    "H": ("Módulos", modulos),
    "I": ("Formatos DIAN", [f"{n} ({c})" for n, c in dian]),
    "J": ("Tipo de cliente", tipos),
}
for col, (titulo, valores) in listas.items():
    c = wl[f"{col}1"]
    c.value = titulo
    c.font = base_font(bold=True, color=WHITE, size=10.5)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center")
    anchos = [len(titulo) + 4] + [len(str(v)) + 2 for v in valores]
    wl.column_dimensions[col].width = max(12, *anchos)
    for i, v in enumerate(valores, start=2):
        wl[f"{col}{i}"] = v
        wl[f"{col}{i}"].font = base_font(size=10)
wl.sheet_view.showGridLines = False
wl.freeze_panes = "A2"

def rng(col, n):
    return f"Listas!${col}$2:${col}${max(n, 1) + 1}"

# =================================================================
# HOJA · CLIENTES  (captura)
# =================================================================
wc = wb.create_sheet("Clientes", index=1)
wc.sheet_view.showGridLines = False

mod_headers = list(modulos)
dian_headers = [f"DIAN · {n} ({c})" for n, c in dian]
headers = [
    ("Razón social *", 30, True),
    ("NIT *", 18, True),
    ("Tipo de cliente *", 16, True),
    ("ERP *", 16, True),
    ("Sector *", 18, True),
    ("Socio (firma) *", 20, True),
    ("Gerente (valida) *", 20, True),
    ("Senior (revisa) *", 20, True),
    ("Staff (ejecuta) * — uno o varios, separar con ;", 34, True),
] + [(h, 15, False) for h in mod_headers] + [(h, 22, False) for h in dian_headers]

N_REQ = 9  # primeras 9 columnas obligatorias (datos + equipo de trabajo)
headers_index = {h[0]: i for i, h in enumerate(headers)}

# Encabezado
for idx, (title, width, req) in enumerate(headers, start=1):
    col = get_column_letter(idx)
    wc.column_dimensions[col].width = width
    c = wc[f"{col}1"]
    c.value = title
    c.font = base_font(bold=True, color=WHITE, size=10)
    c.fill = PatternFill("solid", fgColor=(NAVY if req else NAVY_SOFT))
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
wc.row_dimensions[1].height = 38
wc.freeze_panes = "A2"

# Comentarios de ayuda
staff_col = get_column_letter(headers_index["Staff (ejecuta) * — uno o varios, separar con ;"] + 1)
staff_lista = "; ".join(staffs) if staffs else "(no hay staff activos)"
wc[f"{staff_col}1"].comment = Comment(
    f"Escribe uno o varios staff separados por punto y coma.\nDisponibles: {staff_lista}\nDeben reportar al Senior elegido.",
    "Plantilla",
)
socio_col = get_column_letter(headers_index["Socio (firma) *"] + 1)
wc[f"{socio_col}1"].comment = Comment(
    "Socio responsable de la firma. El Gerente que elijas debe reportar a este Socio.",
    "Plantilla",
)

# Filas de ejemplo (a borrar)
ejemplos = []
if cadenas:
    g, s, st = cadenas[0]
    so = socios[0] if socios else ""
    s1 = st[0] if st else ""
    s2 = "; ".join(st[:2]) if len(st) >= 2 else s1
    ejemplos = [
        ["Inversiones del Pacífico S.A.S", "900.451.227-3", "A",
         (erps[2] if len(erps) > 2 else (erps[0] if erps else "SIESA")),
         (sectores[1] if len(sectores) > 1 else "Comercio"),
         so, g, s, s1,
         "No", "Sí", "Sí", "Sí", "No", "No",   # módulos: AFI,CAR,CXP,ING,INV,NOM
         "Sí", "Sí", "No"],                      # DIAN: IVA, RETE, ICA
        ["Agroindustrias del Cauca Ltda.", "830.118.044-1", "B",
         (erps[3] if len(erps) > 3 else (erps[0] if erps else "SIIGO")),
         (sectores[0] if sectores else "Agroindustria"),
         so, g, s, s2,
         "No", "Sí", "No", "Sí", "Sí", "Sí",
         "Sí", "Sí", "Sí"],
    ]

EX_ROWS = len(ejemplos)
for ri, fila in enumerate(ejemplos, start=2):
    for ci, val in enumerate(fila, start=1):
        cell = wc[f"{get_column_letter(ci)}{ri}"]
        cell.value = val
        cell.font = base_font(size=10, italic=True, color="8A94A3")
        cell.fill = PatternFill("solid", fgColor=EX_FILL)
        cell.border = border
        if ci == 3 or ci >= 10:   # Tipo + módulos/DIAN centrados
            cell.alignment = Alignment(horizontal="center")
nota_col = get_column_letter(len(headers) + 2)
for ri in range(2, 2 + EX_ROWS):
    cell = wc[f"{nota_col}{ri}"]
    cell.value = "← EJEMPLO (borrar)"
    cell.font = base_font(size=9, italic=True, bold=True, color="B22222")

# Filas de captura en blanco con formato
FIRST_DATA = 2 + EX_ROWS
LAST_ROW = 300
for ri in range(FIRST_DATA, LAST_ROW + 1):
    for ci in range(1, len(headers) + 1):
        cell = wc[f"{get_column_letter(ci)}{ri}"]
        cell.border = border
        cell.font = base_font(size=10)
        cell.fill = PatternFill("solid", fgColor=(REQ_FILL if ci <= N_REQ else OPT_FILL))
        if ci == 3 or ci >= 10:
            cell.alignment = Alignment(horizontal="center")

# ---------- Validaciones de datos ----------
APPLY_FROM, APPLY_TO = 2, LAST_ROW

def add_dv(col_idx, formula, strict=True, prompt=None):
    col = get_column_letter(col_idx)
    dv = DataValidation(
        type="list", formula1=formula, allow_blank=True,
        showErrorMessage=strict, showInputMessage=bool(prompt),
        errorStyle="stop" if strict else "warning",
    )
    if strict:
        dv.error = "Valor no válido. Elige uno de la lista."
        dv.errorTitle = "Fuera de catálogo"
    if prompt:
        dv.prompt = prompt
        dv.promptTitle = "Ayuda"
    dv.add(f"{col}{APPLY_FROM}:{col}{APPLY_TO}")
    wc.add_data_validation(dv)

idx = lambda name: headers_index[name] + 1
# Tipo de cliente (estricto A/B/C)
add_dv(idx("Tipo de cliente *"), rng("J", len(tipos)), strict=True,
       prompt="Clasificación del cliente: A, B o C.")
# ERP y Sector: lista con override (warning) — el formulario permite valores nuevos
add_dv(idx("ERP *"), rng("B", len(erps)), strict=False,
       prompt="Elige un ERP existente o escribe uno nuevo.")
add_dv(idx("Sector *"), rng("C", len(sectores)), strict=False,
       prompt="Elige un sector existente o escribe uno nuevo.")
# Socio, Gerente, Senior: estrictos
add_dv(idx("Socio (firma) *"), rng("D", len(socios)), strict=True,
       prompt="Socio responsable de la firma.")
add_dv(idx("Gerente (valida) *"), rng("E", len(gerentes)), strict=True)
add_dv(idx("Senior (revisa) *"), rng("F", len(seniors)), strict=True)
# Módulos + DIAN: Sí/No estricto
for ci in range(10, len(headers) + 1):
    add_dv(ci, rng("A", 2), strict=True, prompt="Sí = activar · No / vacío = no activar")

# Orden final de hojas: Instrucciones, Clientes, Listas
wb.move_sheet("Listas", offset=2)
wb.active = wb.sheetnames.index("Clientes")

out = "/Users/vicbook/Documents/Xentria-apps/Russell Diagnostico/russell-lfm/Plantilla_Importacion_Clientes.xlsx"
wb.save(out)
print("OK", out)
print("socios", socios)
print("cadenas", cadenas)
